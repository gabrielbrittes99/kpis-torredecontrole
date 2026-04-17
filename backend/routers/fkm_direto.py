"""
Abastecimentos Diretos — fora do cartão TruckPag.

Gestores fazem upload de planilha Excel com os abastecimentos realizados
diretamente (em dinheiro, boleto, etc.) e o sistema persiste em
torre.abastecimentos_diretos no DW para compor o FKM.

Endpoints:
  GET  /api/fkm-direto/template        — download da planilha modelo
  POST /api/fkm-direto/upload           — upload, validação e persistência
  GET  /api/fkm-direto/listar           — lista lançamentos do mês
  DELETE /api/fkm-direto/{id}           — remove lançamento
"""
import io
import logging
import os
from datetime import date
from typing import Optional

import openpyxl
import pandas as pd
from fastapi import APIRouter, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/fkm-direto", tags=["fkm-direto"])


def _norm_placa(placa) -> str:
    return str(placa or "").upper().replace("-", "").strip()


def _get_dw():
    from db_dw import get_dw_engine
    return get_dw_engine()


# ── Template ──────────────────────────────────────────────────────────────────

@router.get("/template", summary="Download da planilha modelo para abastecimento direto")
def get_template():
    """Retorna planilha Excel modelo para preenchimento pelos gestores."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Abastecimentos Diretos"

    headers = [
        "Placa", "Data", "Litros", "Valor Total (R$)", "Combustível",
        "Posto / Fornecedor", "Cidade", "UF", "Motorista", "Observação",
    ]
    ws.append(headers)

    # Exemplos de preenchimento na linha 2
    ws.append([
        "BCQ-7B53", "15/04/2026", 120.5, 780.00, "DIESEL S10",
        "Posto Ipiranga", "Curitiba", "PR", "João Silva",
        "Emergência — sem posto credenciado próximo",
    ])

    # Formata cabeçalho
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Ajusta largura das colunas
    col_widths = [12, 12, 10, 16, 16, 24, 16, 6, 20, 40]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=modelo_abastecimento_direto.xlsx"},
    )


# ── Upload ────────────────────────────────────────────────────────────────────

_COLUNAS_ESPERADAS = {
    "placa":      ["placa", "Placa", "PLACA"],
    "data":       ["data", "Data", "DATA"],
    "litros":     ["litros", "Litros", "LITROS"],
    "valor":      ["valor total (r$)", "Valor Total (R$)", "valor", "Valor", "VALOR"],
    "combustivel":["combustível", "Combustível", "combustivel", "COMBUSTIVEL"],
    "posto":      ["posto / fornecedor", "Posto / Fornecedor", "posto", "Posto"],
    "cidade":     ["cidade", "Cidade", "CIDADE"],
    "uf":         ["uf", "UF"],
    "motorista":  ["motorista", "Motorista", "MOTORISTA"],
    "observacao": ["observação", "Observação", "observacao", "Observacao"],
}


def _map_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Mapeia colunas da planilha do gestor para nomes internos."""
    col_lower = {c.lower().strip(): c for c in df.columns}
    rename = {}
    for campo, alternativas in _COLUNAS_ESPERADAS.items():
        for alt in alternativas:
            if alt.lower() in col_lower:
                rename[col_lower[alt.lower()]] = campo
                break
    return df.rename(columns=rename)


def _validar(df: pd.DataFrame) -> list[str]:
    """Retorna lista de erros de validação."""
    erros = []
    obrigatorios = ["placa", "data", "litros", "valor", "combustivel"]
    for col in obrigatorios:
        if col not in df.columns:
            erros.append(f"Coluna obrigatória ausente: '{col}'")

    if erros:
        return erros  # Se faltam colunas, não valida linhas

    for i, row in df.iterrows():
        linha = i + 2  # linha real no Excel (1=header, 2=primeira de dados)
        if not _norm_placa(row.get("placa", "")):
            erros.append(f"Linha {linha}: placa inválida ou vazia")
        try:
            float(row.get("litros", 0) or 0)
        except (ValueError, TypeError):
            erros.append(f"Linha {linha}: litros inválido")
        try:
            float(row.get("valor", 0) or 0)
        except (ValueError, TypeError):
            erros.append(f"Linha {linha}: valor inválido")

    return erros


@router.post("/upload", summary="Upload de planilha de abastecimentos diretos")
async def upload_abastecimentos(
    file: UploadFile = File(...),
    filial: Optional[str] = Query(default=None, description="Filial do gestor que faz upload"),
    upload_por: Optional[str] = Query(default=None, description="Identificação de quem faz upload"),
):
    """
    Recebe planilha Excel com abastecimentos fora TruckPag,
    valida as colunas e persiste em torre.abastecimentos_diretos.
    Retorna preview dos dados inseridos e erros encontrados.
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Arquivo deve ser Excel (.xlsx ou .xls)")

    content = await file.read()
    try:
        df = pd.read_excel(io.BytesIO(content), sheet_name=0)
    except Exception as e:
        raise HTTPException(400, f"Erro ao ler arquivo Excel: {e}")

    df = _map_columns(df)

    # Remove linhas completamente vazias
    df = df.dropna(how="all")
    if df.empty:
        raise HTTPException(400, "Planilha sem dados válidos")

    erros = _validar(df)
    if erros:
        return {"ok": False, "erros": erros, "inseridos": 0, "preview": []}

    # Prepara registros para insert
    hoje = date.today()
    registros = []
    preview = []

    for _, row in df.iterrows():
        placa = _norm_placa(row.get("placa", ""))
        if not placa or len(placa) < 7:
            continue

        # Converte data
        data_raw = row.get("data")
        try:
            if isinstance(data_raw, (date,)):
                data_abastec = data_raw
            else:
                data_abastec = pd.to_datetime(data_raw, dayfirst=True).date()
        except Exception:
            data_abastec = hoje

        ano_mes = f"{data_abastec.year}-{data_abastec.month:02d}"

        litros = float(row.get("litros", 0) or 0)
        valor = float(row.get("valor", 0) or 0)
        combustivel = str(row.get("combustivel", "") or "").strip().upper()
        posto = str(row.get("posto", "") or "").strip()
        cidade = str(row.get("cidade", "") or "").strip()
        uf = str(row.get("uf", "") or "").strip().upper()[:2]
        motorista = str(row.get("motorista", "") or "").strip()
        observacao = str(row.get("observacao", "") or "").strip()

        registros.append({
            "placa": placa,
            "ano_mes": ano_mes,
            "data_abastec": data_abastec,
            "litros": litros,
            "valor": valor,
            "nome_combustivel": combustivel,
            "posto": posto,
            "cidade": cidade,
            "uf": uf,
            "motorista": motorista,
            "observacao": observacao,
            "filial": filial or "",
            "upload_por": upload_por or "",
            "upload_arquivo": file.filename,
        })
        preview.append({
            "placa": placa,
            "ano_mes": ano_mes,
            "data": data_abastec.isoformat(),
            "litros": litros,
            "valor": valor,
            "combustivel": combustivel,
            "posto": posto or None,
        })

    if not registros:
        raise HTTPException(400, "Nenhum registro válido encontrado na planilha")

    # Persiste no DW
    try:
        engine = _get_dw()
        df_insert = pd.DataFrame(registros)
        df_insert.to_sql(
            "abastecimentos_diretos",
            engine,
            schema="torre",
            if_exists="append",
            index=False,
            method="multi",
        )
        logger.info(
            f"fkm_direto: {len(registros)} abastecimentos diretos inseridos "
            f"por {upload_por or 'desconhecido'} ({file.filename})"
        )
    except Exception as e:
        logger.error(f"fkm_direto: falha ao inserir no DW: {e}")
        raise HTTPException(500, f"Erro ao salvar no banco de dados: {e}")

    # Invalida cache FKM para que o próximo GET já traga os dados atualizados
    try:
        from db_fkm_ondemand import refresh_fkm_cache
        refresh_fkm_cache()
    except Exception:
        pass

    return {
        "ok": True,
        "erros": [],
        "inseridos": len(registros),
        "preview": preview,
    }


# ── Listar ────────────────────────────────────────────────────────────────────

@router.get("/listar", summary="Lista abastecimentos diretos do mês")
def listar_abastecimentos(
    ano_mes: Optional[str] = Query(default=None, description="Ex: 2026-04"),
):
    """Lista todos os abastecimentos diretos lançados para o mês informado."""
    if not ano_mes:
        hoje = date.today()
        ano_mes = f"{hoje.year}-{hoje.month:02d}"

    try:
        engine = _get_dw()
        df = pd.read_sql(
            """
            SELECT id, placa, ano_mes, data_abastec, litros, valor,
                   nome_combustivel, posto, cidade, uf, motorista,
                   observacao, filial, upload_por, upload_em, upload_arquivo
            FROM torre.abastecimentos_diretos
            WHERE ano_mes = %(ano_mes)s
            ORDER BY data_abastec, placa
            """,
            engine,
            params={"ano_mes": ano_mes},
        )
    except Exception as e:
        raise HTTPException(500, f"Erro ao consultar banco: {e}")

    return {
        "ano_mes": ano_mes,
        "total_registros": len(df),
        "total_litros": round(float(df["litros"].sum()), 2) if not df.empty else 0,
        "total_valor": round(float(df["valor"].sum()), 2) if not df.empty else 0,
        "registros": df.fillna("").to_dict(orient="records"),
    }


# ── Deletar ───────────────────────────────────────────────────────────────────

@router.delete("/{registro_id}", summary="Remove um lançamento de abastecimento direto")
def deletar_abastecimento(registro_id: int):
    """Remove um registro pelo ID. Invalida o cache FKM após remoção."""
    try:
        engine = _get_dw()
        with engine.connect() as conn:
            result = conn.execute(
                "DELETE FROM torre.abastecimentos_diretos WHERE id = %(id)s RETURNING id",
                {"id": registro_id},
            )
            deleted = result.rowcount
            conn.commit()
    except Exception as e:
        raise HTTPException(500, f"Erro ao deletar: {e}")

    if deleted == 0:
        raise HTTPException(404, f"Registro {registro_id} não encontrado")

    try:
        from db_fkm_ondemand import refresh_fkm_cache
        refresh_fkm_cache()
    except Exception:
        pass

    return {"ok": True, "id": registro_id, "deletado": True}
